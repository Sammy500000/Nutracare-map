#!/usr/bin/env python3
"""Export the scraped hospitals.geojson to an .xlsx workbook:
   - 'Hospitals'      : every hospital with all fields (real coordinates, category, contact…)
   - 'Summary'        : counts by type and by Google category
   - 'By District'    : per-district counts (Government / Private / Maternity / Total)
Output: data/Maharashtra_Hospitals.xlsx
"""
import json, os
import pandas as pd

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
DATA = os.path.join(ROOT, "data")
src = json.load(open(os.path.join(DATA, "hospitals.geojson"), encoding="utf-8"))["features"]

rows = []
for f in src:
    p = f["properties"]
    lon, lat = f["geometry"]["coordinates"]
    rows.append({
        "Name": p.get("name"),
        "Type": p.get("type"),
        "Google Category": p.get("category"),
        "District": p.get("district"),
        "Address": p.get("location"),
        "Pincode": p.get("pincode"),
        "Phone": p.get("phone"),
        "Website": p.get("website"),
        "Rating": p.get("rating"),
        "Reviews": p.get("reviews"),
        "Latitude": lat,
        "Longitude": lon,
        "Place ID": p.get("place_id"),
        "Google Maps Link": p.get("maps_link"),
        "Source": p.get("source"),
    })
df = pd.DataFrame(rows).sort_values(["District", "Type", "Name"]).reset_index(drop=True)

# Summary: by type and by category
type_counts = df["Type"].value_counts().rename_axis("Type").reset_index(name="Count")
cat_counts = df["Google Category"].fillna("(uncategorised)").value_counts().rename_axis("Google Category").reset_index(name="Count")

# By district
piv = (df.pivot_table(index="District", columns="Type", values="Name", aggfunc="count", fill_value=0)
         .reset_index())
piv["Total"] = piv[[c for c in piv.columns if c != "District"]].sum(axis=1)
piv = piv.sort_values("Total", ascending=False)

out = os.path.join(DATA, "Maharashtra_Hospitals.xlsx")
# If the workbook is open (locked), fall back to a non-conflicting name so we never fail.
try:
    open(out, "a").close()
except PermissionError:
    alt = os.path.join(DATA, "Maharashtra_Hospitals_NEW.xlsx")
    print(f"  ! {os.path.basename(out)} is open/locked — writing to {os.path.basename(alt)} instead")
    out = alt
with pd.ExcelWriter(out, engine="openpyxl") as xl:
    df.to_excel(xl, sheet_name="Hospitals", index=False)
    # Summary sheet: type table then category table
    type_counts.to_excel(xl, sheet_name="Summary", index=False, startrow=1)
    cat_counts.to_excel(xl, sheet_name="Summary", index=False, startrow=len(type_counts) + 4)
    piv.to_excel(xl, sheet_name="By District", index=False)

    # light formatting: header bold + autofilter + column widths + frozen header
    from openpyxl.styles import Font
    wb = xl.book
    ws = wb["Hospitals"]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True)
    widths = {"A": 46, "B": 12, "C": 22, "D": 14, "E": 50, "F": 9, "G": 16,
              "H": 34, "I": 8, "J": 9, "K": 11, "L": 11, "M": 30, "N": 44, "O": 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    sm = wb["Summary"]
    sm["A1"] = "Hospitals by Type"; sm["A1"].font = Font(bold=True, size=12)
    sm.cell(row=len(type_counts) + 4, column=1).value = "Hospitals by Google Category"
    sm.cell(row=len(type_counts) + 4, column=1).font = Font(bold=True, size=12)
    for c in wb["By District"][1]:
        c.font = Font(bold=True)

print(f"  wrote {len(df)} hospitals -> {out}")
print("  by type:", dict(zip(type_counts['Type'], type_counts['Count'])))
print(f"  distinct Google categories: {len(cat_counts)}")
